from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import FileResponse
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader
from rapidfuzz import fuzz
import io, csv, re, json

app = FastAPI(title='Stock Statement Compilation Portal')
BASE_DIR = Path(__file__).resolve().parent.parent

@app.get('/', include_in_schema=False)
def home():
    return FileResponse(BASE_DIR / 'index.html')

@app.get('/api/health')
@app.get('/health')
def health():
    return {'ok': True, 'service': 'stock-statement-compiler', 'version': 'sku-match-v5-85-family-safe'}

def num(x):
    if x is None: return 0.0
    s=str(x).strip().replace(',','')
    if s in ('','-','—','–'): return 0.0
    try: return float(s)
    except: return 0.0

def norm_identity(s):
    s=str(s or '').upper().replace('’', "'").replace('&',' AND ')
    s=re.sub(r'[^A-Z0-9]+',' ',s)
    return re.sub(r'\s+',' ',s).strip()

def norm_customer_identity(s):
    s=norm_identity(s)
    # Keep distinctive business words; remove only location / legal-noise words.
    s=re.sub(r'\b(M S|MS|MESSRS|AKOLA|NAGPUR|AMRAVATI|WASHIM|BULDHANA|YAVATMAL|WARDHA|JALGAON|NASHIK|PUNE|AURANGABAD|CHHATRAPATI SAMBHAJINAGAR|HEADQUARTER|HQ|THE)\b',' ',s)
    return re.sub(r'\s+',' ',s).strip()

def identity_score(a,b,customer=False):
    na=norm_customer_identity(a) if customer else norm_identity(a)
    nb=norm_customer_identity(b) if customer else norm_identity(b)
    if not na or not nb: return 0
    if na==nb: return 100
    return round(max(fuzz.ratio(na,nb), fuzz.WRatio(na,nb)))

def clean_candidate(s):
    s=' '.join(str(s or '').replace('\u00a0',' ').split()).strip(' .:-;,|')
    return s

def extract_text_pdf(data):
    reader=PdfReader(io.BytesIO(data))
    chunks=[]
    for page in reader.pages[:5]:
        try: chunks.append(page.extract_text() or '')
        except Exception: pass
    return '\n'.join(chunks)

def extract_pdf_identity(data):
    text=extract_text_pdf(data)
    lines=[clean_candidate(x) for x in text.splitlines() if clean_candidate(x)]
    joined='\n'.join(lines)

    # 1) Explicit labels are the most reliable form.
    stock_patterns=[
        r'(?:STOCKIST|STOCKIST NAME|CUSTOMER|CUSTOMER NAME|PARTY|PARTY NAME|DEALER|DEALER NAME)\s*[:\-]\s*(.+)',
        r'(?:NAME OF STOCKIST|NAME OF CUSTOMER)\s*[:\-]\s*(.+)'
    ]
    hq_patterns=[r'(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*([A-Z][A-Z .&()\-/]+)']
    stock=''; hq=''; raw=''
    for line in lines[:80]:
        for pat in stock_patterns:
            m=re.search(pat,line,re.I)
            if m and len(clean_candidate(m.group(1)))>=3:
                stock=clean_candidate(m.group(1)); raw=line; break
        if stock: break
    for line in lines[:80]:
        for pat in hq_patterns:
            m=re.search(pat,line,re.I)
            if m:
                hq=clean_candidate(m.group(1)); break
        if hq: break

    # 2) Common Alkem stock statement header: CUSTOMER, HQ before report title.
    report_idx=next((i for i,x in enumerate(lines) if re.search(r'Stock\s*&\s*Sales\s*Report',x,re.I)),None)
    candidates=lines[:report_idx] if report_idx is not None else lines[:25]
    # Common Alkem header forms: STOCKIST,HQ on one line or stockist alone on the first line.
    if not stock:
        for line in lines[:12]:
            m=re.match(r'^(.{3,100}?),\s*([A-Za-z][A-Za-z0-9 .&()\-/]{1,40})\s+Stock\s*&\s*Sales\s*Report',line,re.I)
            if m:
                stock=clean_candidate(m.group(1)); hq=hq or clean_candidate(m.group(2)); raw=line; break
    if not stock or not hq:
        for line in candidates:
            if re.search(r'Product\s+Name|ALKEM|Stock\s*&\s*Sales',line,re.I): continue
            m=re.match(r'^(.+?)\s*,\s*([A-Za-z][A-Za-z0-9 .&()\-/]*)$',line)
            if m:
                c=clean_candidate(m.group(1)); q=clean_candidate(m.group(2))
                if len(c)>=3 and len(q)>=2:
                    stock=stock or c; hq=hq or q; raw=raw or line
                    break

    # 3) Some exports put stockist and HQ on adjacent lines.
    if not stock and candidates:
        for i,line in enumerate(candidates[:12]):
            if re.search(r'(STOCK|CUSTOMER|PARTY|DEALER)',line,re.I) and not re.search(r'ALKEM|REPORT',line,re.I):
                nxt=candidates[i+1] if i+1<len(candidates) else ''
                if len(line)>=3:
                    stock=clean_candidate(re.sub(r'^(STOCKIST|CUSTOMER|PARTY|DEALER)\s*(NAME)?\s*[:\-]?\s*','',line,flags=re.I))
                    if not hq and re.fullmatch(r'[A-Za-z .&\-/]{2,30}',nxt): hq=clean_candidate(nxt)
                    raw=line
                    break

    return {'customer':stock,'hq':hq,'source':'PDF text/header' if stock or hq else 'PDF text not detected','raw_header':raw,'raw_text_preview':joined[:2000]}

def excel_values(data, ext):
    if ext in ('.xlsx','.xlsm'):
        wb=load_workbook(io.BytesIO(data),data_only=True,read_only=True)
        for ws in wb.worksheets[:3]:
            vals=[]
            for row in ws.iter_rows(min_row=1,max_row=30,max_col=15,values_only=True):
                vals.append([str(x).strip() if x is not None else '' for x in row])
            yield ws.title, vals
    elif ext=='.xls':
        try:
            import xlrd
        except Exception as e:
            raise ValueError('Legacy .XLS support requires xlrd; please redeploy with the updated requirements.txt') from e
        book=xlrd.open_workbook(file_contents=data,on_demand=True)
        for sh in book.sheets()[:3]:
            vals=[]
            for r in range(min(sh.nrows,40)):
                vals.append([str(sh.cell_value(r,c)).strip() for c in range(min(sh.ncols,20))])
            yield sh.name, vals

def find_labeled(values):
    stock=''; hq=''; raw=''
    if not values:
        return stock,hq,raw

    # Strongest Excel pattern: first row has HQ + CUSTOMER NAME columns.
    for r,row in enumerate(values[:20]):
        labels=[norm_identity(x) for x in row]
        if 'HQ' in labels and ('CUSTOMER NAME' in labels or 'CUSTOMER' in labels):
            hi=labels.index('HQ')
            ci=labels.index('CUSTOMER NAME') if 'CUSTOMER NAME' in labels else labels.index('CUSTOMER')
            for data_row in values[r+1:r+11]:
                if hi<len(data_row) and ci<len(data_row):
                    hv=clean_candidate(data_row[hi]); cv=clean_candidate(data_row[ci])
                    if hv and cv:
                        return cv,hv,' | '.join(clean_candidate(x) for x in data_row if clean_candidate(x))

    # Explicit label/value rows.
    bad_fragments=('SKU NAME','PRODUCT NAME','PTS','SEC UNITS','SEC VALUE','CLO UNITS',
                   'CLO VALUE','SALES','CLOSING','OPENING','ORDER','PENDING','DATE')
    for row in values[:40]:
        clean=[clean_candidate(x) for x in row if clean_candidate(x)]
        line=' '.join(clean)
        if not line: continue
        upper=line.upper()
        if any(b in upper for b in bad_fragments) and not re.search(r'\b(?:HQ|CUSTOMER|STOCKIST|PARTY|DEALER)\s*[:\-]',upper):
            continue

        for pat in [r'(?:STOCKIST|STOCKIST NAME|CUSTOMER|CUSTOMER NAME|PARTY|PARTY NAME|DEALER|DEALER NAME)\s*[:\-]\s*(.+)$']:
            m=re.search(pat,line,re.I)
            if m and not stock:
                val=clean_candidate(m.group(1))
                if val and not any(b in val.upper() for b in bad_fragments):
                    stock=val; raw=line

        for pat in [r'(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*(.+)$']:
            m=re.search(pat,line,re.I)
            if m and not hq:
                val=clean_candidate(m.group(1))
                if val and not any(b in val.upper() for b in bad_fragments):
                    hq=val

    return stock,hq,raw

def extract_text_generic(data, ext):
    if ext not in ('.txt', '.html', '.htm'):
        return ''
    text=data.decode('utf-8-sig','replace')
    if ext in ('.html','.htm'):
        # Preserve line boundaries around common block tags so labels such as
        # Customer Name / HQ remain independently detectable.
        text=re.sub(r'<(br|/p|/div|/tr|/li|/h[1-6])[^>]*>', '\n', text, flags=re.I)
        text=re.sub(r'<script\b[^>]*>.*?</script>', ' ', text, flags=re.I|re.S)
        text=re.sub(r'<style\b[^>]*>.*?</style>', ' ', text, flags=re.I|re.S)
        text=re.sub(r'<[^>]+>', ' ', text)
        text=re.sub(r'&nbsp;', ' ', text, flags=re.I)
        text=re.sub(r'&amp;', '&', text, flags=re.I)
    return '\n'.join(' '.join(line.split()) for line in text.splitlines() if line.strip())

def extract_text_identity_generic(data, ext):
    text=extract_text_generic(data,ext)
    if not text:
        return {'customer':'','hq':'','source':'Text identity not detected','raw_header':''}

    stock=''; hq=''; raw=''
    lines=[clean_candidate(x) for x in text.splitlines() if clean_candidate(x)]

    # Explicit labels, line by line, so one field cannot swallow the next field.
    for line in lines[:100]:
        m=re.match(r'^(?:STOCKIST|STOCKIST NAME|CUSTOMER|CUSTOMER NAME|PARTY|PARTY NAME|DEALER|DEALER NAME)\s*[:\-]\s*(.+)$',line,re.I)
        if m and not stock:
            stock=clean_candidate(m.group(1)); raw=line
        m=re.match(r'^(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*(.+)$',line,re.I)
        if m and not hq:
            hq=clean_candidate(m.group(1))
        if stock and hq: break

    if not stock or not hq:
        # Same-line forms: Customer: X | HQ: Y
        for line in lines[:100]:
            mc=re.search(r'(?:CUSTOMER|CUSTOMER NAME|STOCKIST|STOCKIST NAME|PARTY|PARTY NAME|DEALER|DEALER NAME)\s*[:\-]\s*([^|;,]+)',line,re.I)
            mh=re.search(r'(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*([^|;,]+)',line,re.I)
            if mc and not stock: stock=clean_candidate(mc.group(1)); raw=line
            if mh and not hq: hq=clean_candidate(mh.group(1))
            if stock and hq: break

    if not stock or not hq:
        for line in lines[:100]:
            m=re.match(r'^(.{3,100}?)\s*,\s*([A-Za-z][A-Za-z0-9 .&()\-/]{1,40})$',line)
            if m and not re.search(r'PRODUCT|SKU|SEC|CLO|SALES|REPORT|DATE',line,re.I):
                stock=stock or clean_candidate(m.group(1))
                hq=hq or clean_candidate(m.group(2))
                raw=raw or line
                if stock and hq: break

    return {'customer':stock,'hq':hq,
            'source':'Text/HTML identity' if stock or hq else 'Text identity not detected',
            'raw_header':raw}

def extract_excel_identity(data,ext):
    all_rows=[]
    for sheet,values in excel_values(data,ext):
        # Common Alkem stockist export format: the first non-empty cell is the
        # stockist name, followed by an address line containing the city/HQ.
        early=[]
        for row in values[:10]:
            cells=[clean_candidate(x) for x in row if clean_candidate(x)]
            if cells:
                early.append(cells)
        if early:
            first=early[0][0]
            if (len(first)>=3 and
                not re.search(r'^(PRODUCT|SKU|CUSTOMER|STOCKIST|ALKEM|REPORT|STATEMENT|DATE|STOCK\s*&)', first, re.I)):
                # A single-cell first row is a strong stockist-name signal.
                if len(early[0])==1:
                    address=' '.join(early[1]) if len(early)>1 else ''
                    return {'customer':clean_candidate(first),'hq':'',
                            'source':'Excel stockist header','raw_header':(first+' | '+address).strip(' |'),'sheet':sheet}

        stock,hq,raw=find_labeled(values)
        if stock or hq:
            return {'customer':stock,'hq':hq,'source':'Excel labels','raw_header':raw,'sheet':sheet}
        all_rows.extend(values)

    # Search only early rows for a plausible "customer,HQ" identity line.
    for row in all_rows[:80]:
        cells=[clean_candidate(x) for x in row if clean_candidate(x)]
        if not cells: continue
        line=' '.join(cells)
        if any(b in line.upper() for b in ('SKU NAME','PRODUCT NAME','SEC UNITS','CLO UNITS','SEC VALUE','CLO VALUE')):
            continue
        m=re.match(r'^(.{3,100}?)\s*,\s*([A-Za-z][A-Za-z0-9 .&()\-/]{1,40})$',line)
        if m:
            return {'customer':clean_candidate(m.group(1)),
                    'hq':clean_candidate(m.group(2)),
                    'source':'Excel header','raw_header':line}

    return {'customer':'','hq':'','source':'Excel identity not detected','raw_header':''}

def identify_against_master(identity, master_rows):
    # Deduplicate master identity candidates; duplicate master rows must never
    # change the identity result or create duplicate validation choices.
    unique={}
    for row in master_rows:
        key=(str(row.get('HQ','')).strip().upper(),str(row.get('CUSTOMER NAME','')).strip().upper())
        if key not in unique:
            unique[key]=row
    master_rows=list(unique.values())

    raw_customer=identity.get('customer',''); raw_hq=identity.get('hq','')
    if not raw_customer and not raw_hq:
        return {**identity,'customer':'','hq':'','confidence':0,'match_status':'NOT DETECTED'}
    best=None
    for row in master_rows:
        mc=str(row.get('CUSTOMER NAME','')); mh=str(row.get('HQ',''))
        cs=identity_score(raw_customer,mc,True) if raw_customer else 0
        hs=identity_score(raw_hq,mh,False) if raw_hq else 0
        # If only one side was found, don't invent the other from weak evidence.
        if raw_customer and raw_hq: score=round(cs*0.78+hs*0.22)
        elif raw_customer: score=cs
        else: score=hs
        if best is None or score>best['score']: best={'score':score,'customer':mc,'hq':mh,'cscore':cs,'hscore':hs}
    if not best or best['score']<70:
        return {**identity,'customer':'','hq':'','confidence':best['score'] if best else 0,'match_status':'NOT MATCHED'}
    status='EXACT' if best['score']>=98 and raw_hq and best['hscore']>=95 else ('HIGH CONFIDENCE' if best['score']>=90 else 'REVIEW')
    return {**identity,'customer':best['customer'],'hq':best['hq'],'confidence':round(best['score']),'customer_confidence':round(best['cscore']),'hq_confidence':round(best['hscore']),'match_status':status}

# Existing statement parsing retained.
def sku_norm(s):
    s=str(s or '').upper().replace('’', "'")
    s=s.replace('CILNIREM','CILNIKEM')
    s=re.sub(r'(?<=\d)SMG\b','5MG',s)
    s=s.replace('–','-').replace('—','-')
    s=re.sub(r'\b(?:TABLETS?|TABS?|CAPSULES?|CAPS?|ORAL|ER)\b',' ',s)
    s=re.sub(r"\b\d+(?:X1|X|TAB|TABS|,S|'S|S)\b",' ',s)
    return re.sub(r'[^A-Z0-9./+]+','',s)

def canonical_sku_key(value):
    s=str(value or '').upper().replace('’', "'")
    s=s.replace('CILNIREM','CILNIKEM').replace('–','-').replace('—','-')
    s=re.sub(r'\b(?:TABLETS?|TABS?|CAPSULES?|CAPS?|ORAL|ER)\b',' ',s)
    s=re.sub(r"\b\d+(?:X1|X|TAB|TABS|,S|'S|S)\b",' ',s)
    return re.sub(r'[^A-Z0-9./+]+','',s)

def parse_pdf_text(text):
    rows=[]
    for raw in str(text or '').splitlines():
        line=' '.join(str(raw).replace('\u00a0',' ').split())
        if not line or re.match(r'^[-_=]{5,}',line): continue
        if re.search(r'Product Name\s+Pack|Stock & Sales Report|Last Month Sales|Closing Value|Stk\.Value',line,re.I): continue
        dm=re.search(r'\b\d{2}/\d{2}/\d{2}\b',line)
        if not dm: continue
        pre=line[:dm.start()].strip(); toks=pre.split(); vals=[]; i=len(toks)-1
        while i>=0 and len(vals)<7:
            t=toks[i]
            try: v=0.0 if t in ('-','—','–') else float(t.replace(',',''))
            except: break
            vals.append(v); i-=1
        if len(vals)!=7: continue
        vals.reverse(); prefix=toks[:i+1]; pack_idx=None; pack_re=r"\d+(?:\.\d+)?(?:X1|X|TAB|TABS|,TAB|,S|'S|S|CAP|CAPS)?"
        for j in range(len(prefix)-1,0,-1):
            if re.fullmatch(pack_re,prefix[j],re.I): pack_idx=j; break
        if pack_idx is None: continue
        sku=' '.join(prefix[:pack_idx]).strip()
        if len(sku)<3: continue
        rows.append({'source_sku':sku,'sec':vals[3],'close':vals[4]})
    return rows

def parse_pdf(data):
    reader=PdfReader(io.BytesIO(data)); rows=[]
    for page in reader.pages:
        text=page.extract_text() or ''
        for raw in text.splitlines():
            line=' '.join(str(raw).replace('\u00a0',' ').split())
            if not line or re.match(r'^[-_=]{5,}',line): continue
            if re.search(r'Product Name\s+Pack|Stock & Sales Report|Last Month Sales|Closing Value|Stk\.Value',line,re.I): continue
            dm=re.search(r'\b\d{2}/\d{2}/\d{2}\b',line)
            if not dm: continue
            pre=line[:dm.start()].strip(); toks=pre.split(); vals=[]; i=len(toks)-1
            while i>=0 and len(vals)<7:
                t=toks[i]
                try: v=0.0 if t in ('-','—','–') else float(t.replace(',',''))
                except: break
                vals.append(v); i-=1
            if len(vals)!=7: continue
            vals.reverse(); prefix=toks[:i+1]
            pack_idx=None; pack_re=r"\d+(?:\.\d+)?(?:X1|X|TAB|TABS|,TAB|,S|'S|S|CAP|CAPS)?"
            for j in range(len(prefix)-1,0,-1):
                if re.fullmatch(pack_re,prefix[j],re.I): pack_idx=j; break
            if pack_idx is None: continue
            sku=' '.join(prefix[:pack_idx]).strip()
            if len(sku)<3: continue
            lastsl,open_v,recd,sales,close,order,pend=vals
            rows.append({'source_sku':sku,'sec':sales,'close':close})
    return rows

def parse_csv(data):
    text=data.decode('utf-8-sig','replace'); out=[]; rows=csv.DictReader(io.StringIO(text))
    for r in rows:
        keys={re.sub(r'[^A-Z]','',str(k).upper()):k for k in r}
        def g(*names):
            for n in names:
                if n in keys:return r.get(keys[n])
            return None
        sku=g('PRODUCTNAME','SKU','PRODUCT','ITEM','PRODUCTDESCRIPTION') or ''
        if sku: out.append({'source_sku':str(sku),'sec':num(g('SALES','SECONDARYUNITS','SECONDARY')),'close':num(g('CLOSE','CLOSING','CLOSINGUNITS'))})
    return out

def parse_xlsx(data,ext='.xlsx'):
    for sheet,rows in excel_values(data,ext):
        if not rows: continue
        # Find header row, not necessarily row 1.
        hi=None; headers=None
        for r,row in enumerate(rows[:20]):
            h=[re.sub(r'[^A-Z]','',str(x or '').upper()) for x in row]
            if any(x in h for x in ('PRODUCTNAME','SKU','PRODUCT','ITEM')):
                hi=r; headers=h; break
        if hi is None: continue
        def idx(*names):
            for n in names:
                if n in headers:return headers.index(n)
            return None
        si=idx('PRODUCTNAME','SKU','PRODUCT','ITEM','PRODUCTDESCRIPTION'); sec_i=idx('SALES','SECONDARYUNITS','SECONDARY'); clo_i=idx('CLOSE','CLOSING','CLOSINGUNITS')
        if si is None: continue
        out=[]
        for rr in rows[hi+1:]:
            sku=rr[si] if si<len(rr) else ''
            if sku: out.append({'source_sku':str(sku),'sec':num(rr[sec_i]) if sec_i is not None and sec_i<len(rr) else 0,'close':num(rr[clo_i]) if clo_i is not None and clo_i<len(rr) else 0})
        if out:return out
    raise ValueError('Could not identify SKU/Product column in Excel')

def parse_text_statement(data,ext):
    text=extract_text_generic(data,ext)
    rows=[]
    # Accept CSV-like text where possible.
    if ',' in text and ('PRODUCT' in text.upper() or 'SKU' in text.upper()):
        try:
            return parse_csv(text.encode('utf-8'))
        except Exception:
            pass

    # Generic tab/space-delimited product lines: detect a SKU followed by numeric columns.
    for raw in re.split(r'[\r\n]+', data.decode('utf-8-sig','replace')):
        line=' '.join(raw.replace('\u00a0',' ').split())
        if not line or re.search(r'PRODUCT NAME|SKU NAME|STOCK & SALES|REPORT',line,re.I):
            continue
        nums=re.findall(r'(?<![A-Za-z])[-]?\d+(?:,\d{3})*(?:\.\d+)?',line)
        if len(nums)>=2:
            # Preserve the leading product text; use the last two numeric fields as SEC/CLO.
            first_num=re.search(r'[-]?\d+(?:,\d{3})*(?:\.\d+)?',line)
            if first_num:
                sku=line[:first_num.start()].strip(' ,|\t-')
                if len(sku)>=3:
                    rows.append({'source_sku':sku,'sec':num(nums[-2]),'close':num(nums[-1])})
    return rows

def parse_statement(data,filename):
    ext=Path(filename).suffix.lower()
    if ext=='.pdf': return parse_pdf(data)
    if ext in ('.xlsx','.xlsm','.xls'): return parse_xlsx(data,ext)
    if ext=='.csv': return parse_csv(data)
    if ext in ('.txt','.html','.htm'): return parse_text_statement(data,ext)
    raise ValueError(f'Unsupported file type: {ext}')

AUTO_ACCEPT_THRESHOLD = 85

def sku_family_key(value):
    """Return a conservative product-family key used to prevent cross-SKU-family fuzzy matches."""
    n=norm_identity(value)
    if n.startswith('GLUCORYL MV'):
        return 'GLUCORYL MV'
    if re.match(r'^GLUCORYL M(?:[0-9]| 0 5)(?: |$)', n):
        return 'GLUCORYL M'
    if n.startswith('GLUCORYL MD'):
        return 'GLUCORYL MD'
    if n.startswith('GLUCORYL MP'):
        return 'GLUCORYL MP'
    if n.startswith('GLUCORYL'):
        return 'GLUCORYL'
    return ''

def sku_variant_key(value):
    """Normalize Glucoryl variant ordering so MV1 FORTE == MV FORTE 1, etc."""
    n=norm_identity(value)
    if not n.startswith('GLUCORYL'):
        return ''
    rest=n[len('GLUCORYL'):].strip()
    # Normalize compact product numbers (M1, MV1, MV2) into separate tokens.
    rest=re.sub(r'\b(MV|M)([0-9])\b', r'\1 \2', rest)
    # Only the product-identity portion matters for this guard; packaging and
    # pack-size terms after TABLET/CAP/etc. are ignored.
    rest=re.split(r'\b(?:TABLETS?|TABS?|CAPSULES?|CAPS?|ORAL|ER)\b',rest,1)[0]
    toks=[t for t in rest.split() if t not in {'X1','X'}]
    return ' '.join(sorted(toks))

def match_rows(rows,pool,aliases,threshold):
    # Fixed business rule: >=85% is automatic; <85% is manual.
    out=[]; reviews=[]
    exact_pool=[]
    for idx in range(len(pool)):
        row_master=pool[idx]
        row_key=canonical_sku_key(row_master.get('SKU NAME',''))
        if row_key:
            exact_pool.append((row_key,row_master))

    for ridx in range(len(rows)):
        r=rows[ridx]
        key=canonical_sku_key(r.get('source_sku',''))
        family=sku_family_key(r.get('source_sku',''))
        # For the Glucoryl M/MV families, fuzzy matching must never cross
        # family boundaries (e.g. MV2 FORTE -> M2 FORTE). Other brands keep
        # the existing broad fuzzy matching behaviour.
        candidate_pool=[p for p in pool if (not family or sku_family_key(p.get('SKU NAME',''))==family)]
        if family and candidate_pool:
            variant=sku_variant_key(r.get('source_sku',''))
            variant_candidates=[p for p in candidate_pool if sku_variant_key(p.get('SKU NAME',''))==variant]
            # Exact variant candidates get the first right of refusal. If the
            # statement omits a distinguishing token, retain the family pool
            # and let the 85% threshold decide.
            if variant and variant_candidates:
                candidate_pool=variant_candidates
        if family and not candidate_pool:
            candidate_pool=pool
        chosen=aliases.get(key)
        if chosen:
            m=None
            for j in range(len(pool)):
                if str(pool[j]['SKU NAME'])==chosen:
                    m=pool[j]; break
            if m is not None:
                item={**r,'master_sku':chosen,'suggestion':chosen,'pts':num(m.get('PTS')),'confidence':100,'status':'VALIDATED'}
                out.append(item); continue

        exact_match=None
        for j in range(len(candidate_pool)):
            if canonical_sku_key(candidate_pool[j].get('SKU NAME',''))==key:
                exact_match=candidate_pool[j]; break
        if exact_match is not None:
            master_name=str(exact_match['SKU NAME'])
            item={**r,'master_sku':master_name,'suggestion':master_name,
                  'pts':num(exact_match['PTS']),'confidence':100,'status':'AUTO MATCH'}
            out.append(item); continue

        best=None
        for j in range(len(candidate_pool)):
            p=candidate_pool[j]
            name=str(p['SKU NAME'])
            nk=canonical_sku_key(name)
            if not nk: continue
            score=max(fuzz.ratio(key,nk),fuzz.WRatio(key,nk))
            if best is None or score>best[0]: best=(score,p)
        if best is not None:
            score,p=best
            suggestion=str(p['SKU NAME'])
            accepted=score>=AUTO_ACCEPT_THRESHOLD
            item={**r,'master_sku':suggestion if accepted else '',
                  'suggestion':suggestion,'pts':num(p['PTS']),
                  'confidence':round(score),'status':'AUTO MATCH' if accepted else 'REVIEW'}
        else:
            item={**r,'master_sku':'','suggestion':'','pts':0,'confidence':0,'status':'UNMATCHED'}
        if item['master_sku']: out.append(item)
        reviews.append(item)
    return out,reviews


_MASTER_CACHE = None

def load_master_cache():
    global _MASTER_CACHE
    if _MASTER_CACHE is not None:
        return _MASTER_CACHE
    p = BASE_DIR / 'master-template.xlsx'
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or '').strip() for x in rows[0]]
    out = []
    for row in rows[1:]:
        o = {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers)) if headers[i]}
        if o.get('HQ') and o.get('CUSTOMER NAME') and o.get('SKU NAME'):
            o['PTS'] = num(o.get('PTS'))
            out.append(o)
    # Same logical dedupe key used by the frontend.
    seen = set(); deduped = []
    for r in out:
        k = sku_norm(str(r.get('HQ',''))) + '|' + sku_norm(str(r.get('CUSTOMER NAME',''))) + '|' + sku_norm(str(r.get('SKU NAME','')))
        if k not in seen:
            seen.add(k); deduped.append(r)
    _MASTER_CACHE = deduped
    return _MASTER_CACHE

def raw_filename(request: Request) -> str:
    return request.query_params.get('filename') or request.headers.get('x-filename') or 'statement'

def read_raw_statement(request: Request):
    return request.body()


def identity_from_payload(payload):
    kind=str(payload.get('kind') or '')
    ext=str(payload.get('ext') or '').lower()
    if kind=='text':
        text=str(payload.get('identity_text') or payload.get('text') or '')
        if ext=='.pdf':
            return extract_pdf_identity_from_text(text)
        return extract_text_identity_generic(text.encode('utf-8'),ext)
    if kind=='rows':
        sheets=payload.get('identity_sheets') or []
        values=[]
        for sh in sheets:
            if isinstance(sh,dict): values.extend(sh.get('rows') or [])
        if values:
            if ext in ('.xlsx','.xlsm','.xls','.csv'):
                # Reuse the same labelled Excel/table logic without requiring the file upload.
                stock,hq,raw=find_labeled(values)
                if stock or hq:
                    return {'customer':stock,'hq':hq,'source':'Structured statement identity','raw_header':raw}
                for row in values[:40]:
                    line=' '.join(clean_candidate(x) for x in row if clean_candidate(x))
                    m=re.search(r'(?:CUSTOMER|CUSTOMER NAME|STOCKIST|STOCKIST NAME|PARTY|PARTY NAME)\s*[:\-]\s*([^|;,]+)',line,re.I)
                    q=re.search(r'(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*([^|;,]+)',line,re.I)
                    if m or q:
                        return {'customer':clean_candidate(m.group(1)) if m else '', 'hq':clean_candidate(q.group(1)) if q else '', 'source':'Structured statement identity','raw_header':line}
        return {'customer':'','hq':'','source':'Structured identity not detected','raw_header':''}
    return {'customer':'','hq':'','source':'Identity not detected','raw_header':''}

def extract_pdf_identity_from_text(text):
    lines=[clean_candidate(x) for x in str(text or '').splitlines() if clean_candidate(x)]
    stock='';hq='';raw=''
    stock_patterns=[r'(?:STOCKIST|STOCKIST NAME|CUSTOMER|CUSTOMER NAME|PARTY|PARTY NAME|DEALER|DEALER NAME)\s*[:\-]\s*(.+)',r'(?:NAME OF STOCKIST|NAME OF CUSTOMER)\s*[:\-]\s*(.+)']
    hq_patterns=[r'(?:HQ|H\.Q\.|HEADQUARTER|HEAD QUARTER|LOCATION)\s*[:\-]\s*([A-Z][A-Z .&()\-/]+)']
    for line in lines[:100]:
        for pat in stock_patterns:
            m=re.search(pat,line,re.I)
            if m and len(clean_candidate(m.group(1)))>=3: stock=clean_candidate(m.group(1));raw=line;break
        if stock: break
    for line in lines[:100]:
        for pat in hq_patterns:
            m=re.search(pat,line,re.I)
            if m: hq=clean_candidate(m.group(1));break
        if hq: break
    report_idx=next((i for i,x in enumerate(lines) if re.search(r'Stock\s*&\s*Sales\s*Report',x,re.I)),None)
    candidates=lines[:report_idx] if report_idx is not None else lines[:25]
    if not stock or not hq:
        for line in candidates:
            if re.search(r'Product\s+Name|ALKEM|Stock\s*&\s*Sales',line,re.I): continue
            m=re.match(r'^(.+?)\s*,\s*([A-Za-z][A-Za-z0-9 .&()\-/]*)$',line)
            if m:
                c=clean_candidate(m.group(1));q=clean_candidate(m.group(2))
                if len(c)>=3 and len(q)>=2: stock=stock or c;hq=hq or q;raw=raw or line;break
    if not stock:
        for line in lines[:12]:
            if re.search(r'ALKEM|STOCK\s*&\s*SALES|PRODUCT\s+NAME|REPORT|PAGE|MONTH',line,re.I): continue
            if len(line)>=3 and not re.search(r'^[-_=]+$',line):
                stock=clean_candidate(line); raw=line; break
    return {'customer':stock,'hq':hq,'source':'PDF text identity' if stock or hq else 'PDF identity not detected','raw_header':raw}

def rows_from_payload(payload):
    if str(payload.get('kind') or '')=='rows':
        rows=payload.get('rows') or []
        return [{'source_sku':str(r.get('source_sku') or ''),'sec':num(r.get('sec')),'close':num(r.get('close'))} for r in rows if isinstance(r,dict) and str(r.get('source_sku') or '').strip()]
    if str(payload.get('kind') or '')=='text':
        ext=str(payload.get('ext') or '').lower(); text=str(payload.get('text') or '')
        if ext=='.pdf': return parse_pdf_text(text)
        return parse_text_statement(text.encode('utf-8'),ext)
    return []

@app.post('/api/identify')
@app.post('/identify')
async def identify_endpoint(request: Request):
    """Identify HQ/Stockist from the uploaded statement.

    The frontend sends multipart/form-data. Raw-body mode is retained for
    compatibility with direct API callers.
    """
    try:
        content_type=(request.headers.get('content-type') or '').lower()
        if 'application/json' in content_type:
            payload=await request.json()
            filename=str(payload.get('filename') or 'statement')
            identity=identity_from_payload(payload)
            return identify_against_master(identity,load_master_cache())
        if 'multipart/form-data' in content_type:
            form=await request.form()
            upload=form.get('file')
            if upload is None or not hasattr(upload,'read'):
                raise HTTPException(400,'Missing uploaded statement file.')
            data=await upload.read()
            filename=getattr(upload,'filename',None) or 'statement'
        else:
            data=await request.body()
            filename=raw_filename(request)
        if len(data)>900_000:
            raise HTTPException(413,'Statement is too large for the current Vercel request limit. Please upload a smaller file or split the statement.')
        ext=Path(filename).suffix.lower()
        if ext=='.pdf':
            identity=extract_pdf_identity(data)
        elif ext in ('.xlsx','.xlsm','.xls'):
            identity=extract_excel_identity(data,ext)
        elif ext in ('.txt','.html','.htm'):
            identity=extract_text_identity_generic(data,ext)
        else:
            return {'file':filename,'customer':'','hq':'','confidence':0,'match_status':'UNSUPPORTED','source':'Manual selection'}
        master_rows=load_master_cache()
        return identify_against_master(identity,master_rows)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500,f'{type(e).__name__}: {e}')

@app.post('/api/analyze')
@app.post('/analyze')
async def analyze_endpoint(request: Request):
    """Parse one statement and match its SKUs against the selected Master pool."""
    try:
        content_type=(request.headers.get('content-type') or '').lower()
        hq=''; customer=''; threshold=85; aliases={}; pool_from_client=None; payload=None
        if 'application/json' in content_type:
            body=await request.json(); filename=str(body.get('filename') or 'statement')
            hq=str(body.get('hq') or ''); customer=str(body.get('customer') or '')
            threshold=int(body.get('threshold') or 85); aliases=body.get('aliases_json') or {}; pool_from_client=body.get('pool_json') or None; payload=body.get('payload') or {}
            data=b''
        elif 'multipart/form-data' in content_type:
            form=await request.form()
            upload=form.get('file')
            if upload is None or not hasattr(upload,'read'):
                raise HTTPException(400,'Missing uploaded statement file.')
            data=await upload.read()
            filename=getattr(upload,'filename',None) or 'statement'
            hq=str(form.get('hq') or '')
            customer=str(form.get('customer') or '')
            try: threshold=int(str(form.get('threshold') or 85))
            except Exception: threshold=85
            try: aliases=json.loads(str(form.get('aliases_json') or '{}'))
            except Exception: aliases={}
            try: pool_from_client=json.loads(str(form.get('pool_json') or 'null'))
            except Exception: pool_from_client=None
        else:
            data=await request.body()
            filename=raw_filename(request)
            hq=request.query_params.get('hq','')
            customer=request.query_params.get('customer','')
            try: threshold=int(request.query_params.get('threshold','85') or 85)
            except Exception: threshold=85
            try: aliases=json.loads(request.headers.get('x-aliases','{}') or '{}')
            except Exception: aliases={}
        if len(data)>900_000:
            raise HTTPException(413,'Statement is too large for direct upload. The browser should use structured extraction mode.')
        master=load_master_cache()
        pool=[]
        if pool_from_client:
            # The client pool is used only when it has the required SKU/PTS
            # fields. This keeps matching aligned with the selected HQ/Stockist.
            for r in pool_from_client:
                if isinstance(r,dict) and r.get('SKU_NAME'):
                    pool.append({'HQ':r.get('HQ',hq),'CUSTOMER NAME':r.get('CUSTOMER_NAME',customer),'SKU NAME':r.get('SKU_NAME'),'PTS':num(r.get('PTS'))})
        if not pool and hq and customer:
            pool=[r for r in master if str(r.get('HQ',''))==hq and str(r.get('CUSTOMER NAME',''))==customer]
        if not pool and hq and customer:
            nh=norm_identity(hq); nc=norm_customer_identity(customer)
            pool=[r for r in master if norm_identity(r.get('HQ',''))==nh and norm_customer_identity(r.get('CUSTOMER NAME',''))==nc]
        rows=rows_from_payload(payload) if payload is not None else parse_statement(data,filename)
        matched,reviews=match_rows(rows,pool,aliases,threshold)
        return {'file':filename,'rows':len(rows),
                'sec_total':sum(num(x['sec']) for x in rows),
                'close_total':sum(num(x['close']) for x in rows),
                'matched':matched,'reviews':reviews}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500,f'{type(e).__name__}: {e}')
