from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse
import io, csv, re
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader
from rapidfuzz import fuzz

app = FastAPI(title='Stock Statement Compilation Portal')


def num(x):
    if x is None: return 0.0
    s = str(x).strip().replace(',', '')
    if s in ('', '-', '—', '–'): return 0.0
    try: return float(s)
    except: return 0.0

def sku_norm(s):
    s = str(s or '').upper().replace('’', "'")
    s = s.replace('CILNIREM', 'CILNIKEM')
    s = re.sub(r'(?<=\d)SMG\b', '5MG', s)
    s = s.replace('–','-').replace('—','-')
    s = re.sub(r'\b(?:TABLETS?|TABS?|CAPSULES?|CAPS?|ORAL|ER)\b', ' ', s)
    s = re.sub(r'\b\d+(?:X1|X|TAB|TABS|,S|\'S|S)\b', ' ', s)
    s = re.sub(r'[^A-Z0-9./+]+', '', s)
    return s

def customer_norm(s):
    s = Path(str(s)).stem.upper().replace('.', ' ')
    s = re.sub(r'[-_/]+', ' ', s)
    s = re.sub(r'\b(M/S|MS|MESSRS|AKOLA|NAGPUR|CHA|HEADQUARTER|HQ)\b', ' ', s)
    s = re.sub(r'\b(CORPORATION|CORP|COMPANY|DISTRIBUTOR|DISTRIBUTORS|STOCKIST|AGENCIES|AGENCY|PHARMA|PHARMACEUTICALS)\b', ' ', s)
    s = re.sub(r'\b([A-Z])\s+(?=[A-Z](?:\s|$))', r'\1', s)
    s = re.sub(r'[^A-Z0-9 ]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def parse_pdf(data):
    reader = PdfReader(io.BytesIO(data))
    rows=[]
    for page in reader.pages:
        text = page.extract_text() or ''
        for raw in text.splitlines():
            line = ' '.join(str(raw).replace('\u00a0',' ').split())
            if not line or re.match(r'^[-_=]{5,}', line): continue
            if re.search(r'Product Name\s+Pack|Stock & Sales Report|Last Month Sales|Closing Value|Stk\.Value', line, re.I): continue
            dm=re.search(r'\b\d{2}/\d{2}/\d{2}\b', line)
            if not dm: continue
            pre=line[:dm.start()].strip(); toks=pre.split(); vals=[]; i=len(toks)-1
            while i>=0 and len(vals)<7:
                t=toks[i]
                if t in ('-','—','–'): v=0.0
                else:
                    try: v=float(t.replace(',',''))
                    except: break
                vals.append(v); i-=1
            if len(vals)!=7: continue
            vals.reverse(); prefix=toks[:i+1]
            if not prefix: continue
            pack_idx=None; pack_re=r"\d+(?:\.\d+)?(?:X1|X|TAB|TABS|,TAB|,S|'S|S|CAP|CAPS)?"
            for j in range(len(prefix)-1,0,-1):
                if re.fullmatch(pack_re,prefix[j],re.I): pack_idx=j; break
            if pack_idx is None: continue
            sku=' '.join(prefix[:pack_idx]).strip()
            if len(sku)<3: continue
            lastsl,open_v,recd,sales,close,order,pend=vals
            rows.append({'source_sku':sku,'sec':sales,'close':close})
    return rows

def parse_csv(data):
    text=data.decode('utf-8-sig','replace'); rows=list(csv.DictReader(io.StringIO(text)))
    out=[]
    for r in rows:
        keys={re.sub(r'[^A-Z]','',str(k).upper()):k for k in r}
        sku=r.get(keys.get('PRODUCTNAME','')) or r.get(keys.get('SKU','')) or r.get(keys.get('PRODUCT','')) or ''
        sec=r.get(keys.get('SALES','')) or r.get(keys.get('SECONDARYUNITS','')) or 0
        clo=r.get(keys.get('CLOSE','')) or r.get(keys.get('CLOSING','')) or r.get(keys.get('CLOSINGUNITS','')) or 0
        if sku: out.append({'source_sku':sku,'sec':num(sec),'close':num(clo)})
    return out

def parse_xlsx(data):
    wb=load_workbook(io.BytesIO(data),data_only=True,read_only=True); ws=wb[wb.sheetnames[0]]; rows=list(ws.iter_rows(values_only=True))
    if not rows:return []
    headers=[re.sub(r'[^A-Z]','',str(x or '').upper()) for x in rows[0]]
    def idx(*names):
        for n in names:
            if n in headers:return headers.index(n)
        return None
    si=idx('PRODUCTNAME','SKU','PRODUCT','ITEM','PRODUCTDESCRIPTION'); sec_i=idx('SALES','SECONDARYUNITS','SECONDARY'); clo_i=idx('CLOSE','CLOSING','CLOSINGUNITS')
    if si is None: raise ValueError('Could not identify SKU/Product column in Excel')
    out=[]
    for rr in rows[1:]:
        sku=rr[si] if si<len(rr) else ''
        if sku: out.append({'source_sku':str(sku),'sec':num(rr[sec_i]) if sec_i is not None and sec_i<len(rr) else 0,'close':num(rr[clo_i]) if clo_i is not None and clo_i<len(rr) else 0})
    return out

def parse_txt(data):
    text=data.decode('utf-8-sig','replace'); out=[]
    for line in text.splitlines():
        parts=re.split(r'\t|,|\s{2,}',line.strip())
        if len(parts)>=3:
            try: sec=num(parts[-2]); clo=num(parts[-1]); sku=' '.join(parts[:-2]); out.append({'source_sku':sku,'sec':sec,'close':clo})
            except: pass
    return out

def parse_statement(data, filename):
    ext=Path(filename).suffix.lower()
    if ext=='.pdf': return parse_pdf(data)
    if ext in ('.xlsx','.xls'): return parse_xlsx(data)
    if ext=='.csv': return parse_csv(data)
    if ext=='.txt': return parse_txt(data)
    raise ValueError(f'Unsupported file type: {ext}')

def match_rows(rows,pool,aliases,threshold):
    out=[]; reviews=[]
    for r in rows:
        key=sku_norm(r['source_sku'])
        if key in aliases:
            chosen=aliases[key]; m=next((p for p in pool if p['SKU_NAME']==chosen),None)
            if m:
                out.append({**r,'master_sku':m['SKU_NAME'],'pts':num(m['PTS']),'confidence':100,'status':'VALIDATED'}); continue
        best=None
        for p in pool:
            score=max(fuzz.ratio(key,sku_norm(p['SKU_NAME'])),fuzz.WRatio(key,sku_norm(p['SKU_NAME'])))
            if best is None or score>best[0]: best=(score,p)
        if best:
            score,p=best; item={**r,'master_sku':p['SKU_NAME'] if score>=threshold else '','suggestion':p['SKU_NAME'],'pts':num(p['PTS']),'confidence':round(score),'status':'AUTO MATCH' if score>=threshold else 'REVIEW'}
        else:
            item={**r,'master_sku':'','suggestion':'','pts':0,'confidence':0,'status':'UNMATCHED'}
        if item['master_sku']:
            out.append(item)
        else:
            reviews.append(item)
    return out,reviews

@app.post("/")
async def analyze_endpoint(file: UploadFile = File(...), pool_json: str = Form(...), aliases_json: str = Form('{}'), threshold: int = Form(82)):
    try:
        import json
        data = await file.read()
        if len(data) > 4_450_000:
            raise HTTPException(413, 'Statement exceeds Vercel request limit (~4.4 MB).')
        pool = json.loads(pool_json)
        aliases = json.loads(aliases_json or '{}')
        rows = parse_statement(data, file.filename)
        matched, reviews = match_rows(rows, pool, aliases, threshold)
        return {
            'file': file.filename,
            'rows': len(rows),
            'sec_total': sum(num(x['sec']) for x in rows),
            'close_total': sum(num(x['close']) for x in rows),
            'matched': matched,
            'reviews': reviews
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'{type(e).__name__}: {e}')
